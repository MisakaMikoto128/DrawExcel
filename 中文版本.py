from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import cv2 as cv
def bgr2hex(bgr):
    (b, g, r) = (bgr)
    return hex(r << 16 | g << 8 | b).replace('0x','').zfill(6)
def 在Excel中画(img):
    W, H, C = img.shape
    for w in range(W):
        sht.column_dimensions[get_column_letter(w + 1)].width = 2.15  #改列宽
        for h in range(H):
            sht.cell(w + 1, h + 1).fill = PatternFill("solid", fgColor=bgr2hex(img[w, h]))#给格子填充颜色
img = cv.imread("Alice2.jpg")#读取图片
wb = Workbook()#新建一个Workbook对象
sht = wb.worksheets[0]#打开第一张表单
# size = img.shape#获取图片大小（长，宽，通道数）
# img = cv.resize(img,(int(size[0]*0.5),int(size[1]*0.5)),cv.INTER_LINEAR)#缩放图片，图片太大会出问题
在Excel中画(img)
wb.save("Alice.xlsx")#保存